
# imports
from tkinter import *
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
import datetime
from dataclasses import dataclass

# setup

# creates a spreadsheet/opens the spreadsheet?


path = "D:\\Python\\other python scripts\\task_testing_2.xlsx"



taskbook = load_workbook(filename=path)


ws = taskbook.active
ws1 = taskbook.create_sheet("tasks testing 2")



# theoretically editable list of task attributes
task_attributes = ["number", "name", "description"]



# set up tkinter
window = Tk()
window.mainloop()



'''
# set up class for messing with the current task
class task:
    def __init__(self):
        # do we need anything?
    def edit(parameter, new_value):
        # need some way to check which column each parameter is located
'''

'''

each parameter gets its own column
when adding a new parameter, a new column is simply appended
tasks that don't have a value for a parameter simply don't have a value
'''
