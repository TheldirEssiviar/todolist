'''
# imports
from tkinter import *
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook, Workbook
import datetime
from dataclasses import dataclass

# setup
'''
# creates a spreadsheet/opens the spreadsheet?

# workbook_cells.py

# workbook_cells.py

from openpyxl import load_workbook

def get_cell_info(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    print(sheet)
    print(f'The title of the Worksheet is: {sheet.title}')
    print(f'The value of {sheet["A2"].value=}')
    print(f'The value of {sheet["A3"].value=}')
    cell = sheet['B3']
    print(f'{cell.value=}')

if __name__ == '__main__':
    get_cell_info('books.xlsx')
    
text = input()


'''

path = "D:\\Python\\other python scripts\\task_testing_2.xlsx"



taskbook = load_workbook("tasks_testing_2.xlsx")


ws = taskbook.active()



# theoretically editable list of task attributes
#task_attributes = ["number", "name", "description"]



# set up tkinter
window = Tk()
window.mainloop()
'''


'''
# set up class for messing with the current task
class task:
    def __init__(self):
        # do we need anything?
    def edit(parameter, new_value):
        # need some way to check which column each parameter is located
'''

'''

each parameter gets its own column
when adding a new parameter, a new column is simply appended
tasks that don't have a value for a parameter simply don't have a value
'''
